import streamlit as st
from openpyxl import load_workbook
import re
from io import BytesIO

st.set_page_config(
    page_title="Integração Folha ADM x CONT",
    layout="wide"
)

st.title("Integração Folha ADM x CONT")

st.write("""
Este sistema:

1. Lê os centros de custo automaticamente do ADM
2. Lê os eventos dos blocos A/D e F/I
3. Armazena os valores por Centro de Custo
4. Procura os eventos na planilha CONT
5. Soma os valores do respectivo Centro de Custo
6. Preenche a coluna H (VLR LANÇAMENTO)
""")

# =========================
# UPLOADS
# =========================
arquivo_adm = st.file_uploader(
    "Selecione a planilha Folha - ADM",
    type=["xlsx"]
)

arquivo_cont = st.file_uploader(
    "Selecione a planilha FOLHA - CONT",
    type=["xlsx"]
)

# =========================
# PROCESSAMENTO
# =========================
if arquivo_adm and arquivo_cont:

    with st.spinner("Processando arquivos..."):

        # =========================
        # LEITURA ADM
        # =========================
        wb_adm = load_workbook(
            arquivo_adm,
            data_only=True
        )

        ws_adm = wb_adm[wb_adm.sheetnames[0]]

        # Estrutura:
        # eventos[cc][codigo] = valor
        eventos = {}

        centro_custo_atual = None

        for row in range(1, ws_adm.max_row + 1):

            texto_a = ws_adm[f"A{row}"].value

            # =====================================
            # IDENTIFICA NOVO CENTRO DE CUSTO
            # =====================================
            if texto_a:

                texto_upper = str(texto_a).upper()

                if "RELATORIO PARA CONTABILIDADE" in texto_upper:

                    match = re.search(
                        r"OBRA\s+(\d+)",
                        texto_upper
                    )

                    if match:

                        centro_custo_atual = int(
                            match.group(1)
                        )

                        if centro_custo_atual not in eventos:

                            eventos[
                                centro_custo_atual
                            ] = {}

                    continue

            # Se ainda não encontrou um CC
            if centro_custo_atual is None:
                continue

            # =====================================
            # BLOCO A / D
            # =====================================
            evento_1 = ws_adm[f"A{row}"].value
            valor_1 = ws_adm[f"D{row}"].value

            if evento_1 not in [None, ""]:

                try:

                    codigo = int(float(evento_1))
                    valor = float(valor_1 or 0)

                    eventos[centro_custo_atual][codigo] = (
                        eventos[centro_custo_atual].get(
                            codigo,
                            0
                        ) + valor
                    )

                except:
                    pass

            # =====================================
            # BLOCO F / I
            # =====================================
            evento_2 = ws_adm[f"F{row}"].value
            valor_2 = ws_adm[f"I{row}"].value

            if evento_2 not in [None, ""]:

                try:

                    codigo = int(float(evento_2))
                    valor = float(valor_2 or 0)

                    eventos[centro_custo_atual][codigo] = (
                        eventos[centro_custo_atual].get(
                            codigo,
                            0
                        ) + valor
                    )

                except:
                    pass

        # =========================
        # LEITURA CONT
        # =========================
        wb_cont = load_workbook(
            arquivo_cont
        )

        ws_cont = wb_cont["ADMIN"]

        regex_codigos = re.compile(r"\d+")

        # =========================
        # PREENCHIMENTO
        # =========================
        for row in range(5, ws_cont.max_row + 1):

            conta = ws_cont[f"B{row}"].value
            cc = ws_cont[f"C{row}"].value

            if not conta:
                continue

            try:
                cc = int(float(cc))
            except:
                continue

            codigos = regex_codigos.findall(
                str(conta)
            )

            soma = 0

            if cc in eventos:

                for codigo in codigos:

                    soma += eventos[cc].get(
                        int(codigo),
                        0
                    )

            celula = ws_cont[f"H{row}"]

            celula.value = soma

            celula.number_format = '#,##0.00'

        # =========================
        # DOWNLOAD
        # =========================
        output = BytesIO()

        wb_cont.save(output)

        output.seek(0)

        st.success(
            "Arquivo processado com sucesso!"
        )

        st.download_button(
            label="📥 Baixar Arquivo Preenchido",
            data=output,
            file_name="FOLHA_02_CONT_PREENCHIDA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
